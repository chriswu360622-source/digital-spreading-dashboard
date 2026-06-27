from __future__ import annotations

import json
from datetime import datetime, date, time, timedelta
from functools import cmp_to_key
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parents[1]
SOURCE_DIR = WORKSPACE / "Digital Spreading Dashboard Input"
OUTPUT_JSON = ROOT / "data" / "dashboard-data.json"
OUTPUT_JS = ROOT / "data" / "dashboard-data.js"


def find_source_workbook() -> Path:
    candidates = [
        path
        for path in SOURCE_DIR.rglob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    if candidates:
        return max(candidates, key=lambda path: path.stat().st_mtime)
    raise FileNotFoundError(
        f"No Excel workbook found in {SOURCE_DIR}. Place a .xlsx file there."
    )


def excel_serial_to_datetime(value: float) -> datetime:
    return datetime(1899, 12, 30) + timedelta(days=float(value))


def normalize_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return excel_serial_to_datetime(value).date().isoformat()
    text = str(value).strip().replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    parts = text.split("-")
    if len(parts) == 3:
        try:
            y, m, d = [int(p) for p in parts]
            return date(y, m, d).isoformat()
        except ValueError:
            return text
    return text


def normalize_datetime(value: Any, fallback_date: str | None = None) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date) and not isinstance(value, datetime):
        return datetime.combine(value, time()).isoformat()
    if isinstance(value, time):
        base = date.fromisoformat(fallback_date) if fallback_date else date(1899, 12, 30)
        return datetime.combine(base, value).isoformat()
    if isinstance(value, (int, float)):
        return excel_serial_to_datetime(value).isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%H:%M:%S", "%H:%M"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt.startswith("%H") and fallback_date:
                parsed = datetime.combine(date.fromisoformat(fallback_date), parsed.time())
            return parsed.isoformat()
        except ValueError:
            pass
    return text


def to_datetime(value: Any, fallback_date: str | None = None) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date) and not isinstance(value, datetime):
        return datetime.combine(value, time())
    if isinstance(value, time):
        base = date.fromisoformat(fallback_date) if fallback_date else date(1899, 12, 30)
        return datetime.combine(base, value)
    if isinstance(value, (int, float)):
        return excel_serial_to_datetime(value)
    text_value = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%H:%M:%S", "%H:%M"):
        try:
            parsed = datetime.strptime(text_value, fmt)
            if fmt.startswith("%H") and fallback_date:
                parsed = datetime.combine(date.fromisoformat(fallback_date), parsed.time())
            return parsed
        except ValueError:
            pass
    return None


def elapsed_hours(start_value: Any, end_value: Any, fallback_date: str | None = None) -> float:
    start_dt = to_datetime(start_value, fallback_date)
    end_dt = to_datetime(end_value, fallback_date)
    if start_dt and end_dt:
        delta = (end_dt - start_dt).total_seconds()
        if delta >= 0:
            return delta / 3600
    return 0.0


def time_to_hours(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, time):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, datetime):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, (int, float)):
        return float(value) * 24
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.hour + parsed.minute / 60 + parsed.second / 3600
        except ValueError:
            pass
    return 0.0


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def rows_from_sheet(wb, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[sheet_name]
    headers = [ws.cell(1, col).value or f"COL{col}" for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        row = {str(header): ws.cell(row_idx, col).value for col, header in enumerate(headers, start=1)}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def pick_sheet_name(wb, preferred: list[str], fallback_contains: list[str] | None = None) -> str:
    available = set(wb.sheetnames)
    for name in preferred:
        if name in available:
            return name
    if fallback_contains:
        for name in wb.sheetnames:
            lowered = name.casefold()
            if any(token.casefold() in lowered for token in fallback_contains):
                return name
    raise KeyError(f"None of the expected sheets exist. Available sheets: {', '.join(wb.sheetnames)}")


def build() -> None:
    source_xlsx = find_source_workbook()
    wb = load_workbook(source_xlsx, data_only=True)
    summary_sheet = pick_sheet_name(wb, ["Summary"], ["summary"])
    detail_sheet = pick_sheet_name(wb, ["Detail"], ["detail"])
    summary_rows = rows_from_sheet(wb, summary_sheet)
    detail_rows = rows_from_sheet(wb, detail_sheet)
    table_rows = []
    for candidate in ["Sheet3", "data", "Data"]:
        if candidate in wb.sheetnames:
            table_rows = rows_from_sheet(wb, candidate)
            break

    summary: list[dict[str, Any]] = []
    summary_by_cutref: dict[str, dict[str, Any]] = {}
    for row in summary_rows:
        spreading_date = normalize_date(row.get("Spreading Date"))
        cutref = text(row.get("CutRef"))
        start_time = normalize_datetime(row.get("Start Time"), spreading_date)
        end_time = normalize_datetime(row.get("End Time"), spreading_date)
        activity_date = (start_time or spreading_date or "")[:10] or spreading_date
        item = {
            "factory": text(row.get("Factory")),
            "cutRef": cutref,
            "status": text(row.get("Spreading Status")),
            "spreadingRef": text(row.get("Spreading Ref")),
            "spreadingDate": spreading_date,
            "activityDate": activity_date,
            "cutCell": text(row.get("Cut Cell")),
            "cutNo": text(row.get("Cut#")),
            "markerName": text(row.get("Marker Name")),
            "markerNo": text(row.get("Marker No.")),
            "markerLengthYard": number(row.get("Marker Length (Yard)")),
            "fabricCombo": text(row.get("Fabric Combo")),
            "article": text(row.get("Article")),
            "color": text(row.get("Color")),
            "size": text(row.get("Size")),
            "layers": number(row.get("Layers")),
            "totalConsYard": number(row.get("Total Cons. (Yard)")),
            "sp": text(row.get("SP")),
            "subSp": text(row.get("Sub SP")),
            "spreadingTable": text(row.get("Spreading Table")),
            "layerSpread": number(row.get("Layer Spread")),
            "totalYardsSpread": number(row.get("Total Yards (Spread)")),
            "balYards": number(row.get("Bal. Yards")),
            "remarkSpreading": text(row.get("Remark (Spreading)")),
            "spreader": text(row.get("Spreader")),
            "startTime": start_time,
            "endTime": end_time,
            "spreadingTimeHours": elapsed_hours(row.get("Start Time"), row.get("End Time"), spreading_date),
            "spreadingTimeSeconds": elapsed_hours(row.get("Start Time"), row.get("End Time"), spreading_date) * 3600,
            "spreaderName": text(row.get("Spreader Name")),
        }
        summary.append(item)
        summary_by_cutref[cutref] = item

    detail: list[dict[str, Any]] = []
    for row in detail_rows:
        cutref = text(row.get("CutRef"))
        parent = summary_by_cutref.get(cutref, {})
        spreading_date = parent.get("spreadingDate")
        start_time = normalize_datetime(row.get("Start Time"), spreading_date)
        end_time = normalize_datetime(row.get("End Time"), spreading_date)
        activity_date = (start_time or spreading_date or "")[:10] or spreading_date
        item = {
            "factory": text(row.get("Factory")),
            "cutRef": cutref,
            "status": text(row.get("Spreading Status")),
            "summaryStatus": parent.get("status", text(row.get("Spreading Status"))),
            "summaryDate": spreading_date,
            "activityDate": activity_date,
            "spreadingRef": text(row.get("Spreading Ref")),
            "cutCell": text(row.get("Cut Cell")),
            "cutNo": text(row.get("Cut#")),
            "subCutNo": text(row.get("Sub CutNo")),
            "markerName": text(row.get("Marker Name")),
            "markerNo": text(row.get("Marker No.")),
            "markerLengthYard": number(row.get("Marker Length (Yard)")),
            "fabricCombo": text(row.get("Fabric Combo")),
            "article": text(row.get("Article")),
            "color": text(row.get("Color")),
            "size": text(row.get("Size")),
            "layers": number(row.get("Layers")),
            "totalConsYard": number(row.get("Total Cons. (Yard)")),
            "sp": text(row.get("SP")) or parent.get("sp", ""),
            "subSp": text(row.get("Sub SP")) or parent.get("subSp", ""),
            "spreadingTable": text(row.get("Spreading Table")),
            "seq": text(row.get("Seq")),
            "roll": text(row.get("Roll")),
            "dyelot": text(row.get("Dyelot")),
            "fabricTone": text(row.get("Fabric Tone")),
            "ticketRemainYards": number(row.get("Ticket /Remain Yards")),
            "layerSpread": number(row.get("Layer Spread")),
            "totalYardsSpread": number(row.get("Total Yards (Spread)")),
            "mergeFabricYard": number(row.get("Merge Fabric (Yard)")),
            "useCutendsYard": number(row.get("Use Cutends (Yard)")),
            "damageYard": number(row.get("Damage (Yard)")),
            "remainYards": number(row.get("Remain Yards")),
            "oriCutendsYard": number(row.get("Ori Cutends (Yard)")),
            "varianceYard": number(row.get("Variance (Yard)")),
            "startTime": start_time,
            "endTime": end_time,
            "spreadingTimeHours": elapsed_hours(row.get("Start Time"), row.get("End Time"), spreading_date),
            "spreader": parent.get("spreader", ""),
            "spreaderName": parent.get("spreaderName", ""),
            "remarkSpreading": parent.get("remarkSpreading", ""),
        }
        detail.append(item)

    spreading_tables = []
    if table_rows:
        for row in table_rows:
            table = text(row.get("Spreading Table"))
            if table:
                spreading_tables.append(
                    {
                        "spreadingTable": table,
                        "idNumber": text(row.get("ID Number")),
                        "name": text(row.get("Name")),
                        "note": text(row.get("COL4")),
                    }
                )
    else:
        spreading_tables = [
            {"spreadingTable": table}
            for table in sorted({row["spreadingTable"] for row in summary if row["spreadingTable"]}, key=cmp_to_key(tableSort))
        ]

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceWorkbook": str(source_xlsx),
        "sourceFolder": str(SOURCE_DIR),
        "defaults": {
            "hourlyTarget": 450,
            "normalShiftHours": 7.5,
            "overtimeStart": "16:30",
            "overtimeRoundingMinutes": 30,
            "defaultStatus": "Finished",
        },
        "summary": summary,
        "detail": detail,
        "spreadingTables": spreading_tables,
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    json_text = json.dumps(payload, ensure_ascii=False, indent=2)
    OUTPUT_JSON.write_text(json_text, encoding="utf-8")
    OUTPUT_JS.write_text("window.DIGITAL_SPREADING_DATA = " + json_text + ";\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")
    print(f"Wrote {OUTPUT_JS}")
    print(f"Summary rows: {len(summary)}")
    print(f"Detail rows: {len(detail)}")


if __name__ == "__main__":
    build()
